"""Weekly schedule .xlsx export (member-weekly-schedule-xlsx-export, 2026-07-24).

Builds the two-worksheet workbook ("Weekly Schedule", "Weekly Summary") for
GET /api/member-schedules/{member_key}/reports/weekly/export
(backend/routers/member_schedules.py). This module owns row shaping,
sorting, and openpyxl formatting only — it never queries the database and
never recomputes Schedule Summary aggregation; the aggregate dict passed
into build_weekly_schedule_workbook() must come from the existing
_aggregate_schedule_period() (member_schedules.py), unchanged.

This is a point-in-time export only. Nothing here writes to the database,
and the generated file has no macros, external links, or upload/sync
affordance — opening or editing it in Excel/Google Sheets never reaches
back into Management AIOS.
"""

from datetime import date as date_type, datetime
from io import BytesIO
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo

from backend.config import SCHEDULE_TIMEZONE
from backend.routers import leave_logic
from backend.time_utils import derive_task_outcome

_COLOMBO = ZoneInfo(SCHEDULE_TIMEZONE)

DAY_NAMES = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")

WEEKLY_SCHEDULE_HEADERS = (
    "Date", "Day", "Start Time", "End Time", "Title / Leave Type",
    "Priority", "Scheduled/Unscheduled Category", "Outcome",
    "Outcome Reason", "Notes", "Created At", "Updated At",
)

# Column widths tuned for the header list above — wide enough that the
# common case (a task title, a leave note) is readable without manual
# resizing once opened in Excel/Google Sheets; Title/Outcome Reason/Notes
# also get wrap_text (see _write_weekly_schedule_sheet) rather than being
# widened further, since those three fields can run to 120-250 characters.
WEEKLY_SCHEDULE_COLUMN_WIDTHS = (12, 11, 10, 10, 34, 10, 24, 13, 30, 30, 17, 17)

# Same display wording the frontend already uses for a Leave calendar chip
# (web-view/js/calendar/core.js LEAVE_TYPE_DISPLAY_LABEL) — kept as a
# separate constant here (not imported, no JS<->Python shared module exists
# in this codebase) but the values themselves must never drift from it.
LEAVE_TYPE_EXPORT_LABEL = {
    "Short Leave": "Short Leave",
    "Half-Day First": "First-Half Leave",
    "Half-Day Second": "Second-Half Leave",
    "Full-Day": "Full-Day Leave",
    "Multi-Day": "Multi-Day Leave",
}

POINT_IN_TIME_NOTICE = (
    "This workbook is a point-in-time export from Management AIOS. Update "
    "Tasks and Leave inside Management AIOS, not in this file."
)

EXPORT_SOURCE_LABEL = "Management AIOS"


def build_export_filename(member_key: str, week_start: date_type, week_end: date_type) -> str:
    """DD-MM-YYYY_to_DD-MM-YYYY_<member-key>_weekly_schedule.xlsx — the one
    place this filename is assembled, so the Content-Disposition header
    (member_schedules.py) can never disagree with what a caller re-derives
    from the same (member_key, week_start, week_end)."""
    return (
        week_start.strftime("%d-%m-%Y") + "_to_" + week_end.strftime("%d-%m-%Y")
        + "_" + member_key + "_weekly_schedule.xlsx"
    )


def _to_colombo_naive(dt: Optional[datetime]) -> Optional[datetime]:
    """Converts an authoritative UTC-aware timestamp to a naive Asia/Colombo
    datetime for spreadsheet display only. Never written back to the
    database or to any request payload — display-only, matching
    core.js formatTaskTimestamp()'s existing Asia/Colombo convention."""
    if dt is None:
        return None
    return dt.astimezone(_COLOMBO).replace(tzinfo=None)


def _format_duration(minutes: Optional[int]) -> str:
    """Same 'Xh Ym' formatting as the frontend's formatDuration()
    (web-view/js/calendar/core.js) — this only formats a minutes value the
    backend's own _aggregate_schedule_period already computed; no duration
    is calculated here."""
    total = max(0, int(minutes or 0))
    hours, mins = divmod(total, 60)
    return str(hours) + "h " + str(mins) + "m"


def task_row(event) -> dict:
    """One Weekly Schedule row for an active MemberScheduleEvent. `outcome`
    is the same derived display value (Pending/Completed/Uncompleted/No
    response) MemberScheduleEventOut returns via derive_task_outcome — never
    the raw stored column read directly, so a caller of this export always
    sees exactly what the API itself would show for this task today."""
    outcome_status, _locked = derive_task_outcome(event.event_date, event.outcome)
    return {
        "date": event.event_date,
        "day": DAY_NAMES[event.event_date.weekday()],
        "start_time": event.start_time,
        "end_time": event.end_time,
        "title": event.title,
        "priority": event.priority,
        "category": event.category,
        "outcome": outcome_status,
        "outcome_reason": event.outcome_reason if outcome_status == "Uncompleted" else None,
        "notes": event.notes,
        "created_at": _to_colombo_naive(event.created_at),
        "updated_at": _to_colombo_naive(event.updated_at),
    }


def leave_rows_for_record(record, week_start: date_type, week_end: date_type) -> List[dict]:
    """Weekly Schedule row(s) for one active MemberLeaveRecord, clamped to
    [week_start, week_end].

    MULTI-DAY LEAVE REPRESENTATION (Step 9 — documented choice): one row
    per covered date, using the exact same Monday-Friday weekday expansion
    (leave_logic.expand_weekdays) the backend's own leave-deduction
    aggregation and the frontend's calendar-chip placement
    (expandWeekdaysClientSide) already use — so a Multi-Day row here can
    never appear on a weekend the calendar itself never shows it on, and
    can never disagree with which dates are "covered" elsewhere in this
    codebase. The leave's own range is clamped to the export week FIRST,
    then expanded — clamping does not change which of the remaining dates
    are weekdays, so this is equivalent to expanding the full range and
    filtering, without ever generating a date outside the requested week.

    Every other leave type (Short Leave, Half-Day First, Half-Day Second,
    Full-Day) is single-day by construction (DB CHECK constraint:
    end_date == start_date) and always produces exactly one row, since the
    caller only ever passes records whose [start_date, end_date] already
    overlaps [week_start, week_end]."""
    label = LEAVE_TYPE_EXPORT_LABEL.get(record.leave_type, record.leave_type)

    if record.leave_type == "Multi-Day":
        clamped_start = max(record.start_date, week_start)
        clamped_end = min(record.end_date, week_end)
        covered_dates = leave_logic.expand_weekdays(clamped_start, clamped_end)
    else:
        covered_dates = (
            [record.start_date] if week_start <= record.start_date <= week_end else []
        )

    created_at = _to_colombo_naive(record.created_at)
    updated_at = _to_colombo_naive(record.updated_at)

    rows = []
    for covered_date in covered_dates:
        rows.append({
            "date": covered_date,
            "day": DAY_NAMES[covered_date.weekday()],
            # Only Short Leave ever carries a real start_time/end_time
            # (DB CHECK constraint forces both NULL for every other leave
            # type) — every other leave type's Start Time/End Time is left
            # blank here rather than substituting the display-only
            # 08:30-13:00/13:30-18:00/08:30-18:00 windows core.js uses
            # purely for Week/Day grid positioning (never invent a stored
            # value that does not exist for a field that does not apply).
            "start_time": record.start_time,
            "end_time": record.end_time,
            "title": label,
            "priority": None,
            "category": None,
            "outcome": None,
            "outcome_reason": None,
            "notes": record.purpose,
            "created_at": created_at,
            "updated_at": updated_at,
        })
    return rows


def _row_sort_key(row: dict):
    """Date, Start Time, End Time, Title/Leave Type (Step 10) — blank
    Start/End Time sort as "" (before any "HH:MM" string), which keeps
    untimed rows grouped first within a date rather than interleaved
    unpredictably; Python's sort is stable, so equal keys preserve
    insertion order."""
    start = row["start_time"].strftime("%H:%M") if row["start_time"] else ""
    end = row["end_time"].strftime("%H:%M") if row["end_time"] else ""
    return (row["date"], start, end, (row["title"] or "").casefold())


def build_weekly_schedule_rows(tasks, leave_records, week_start: date_type, week_end: date_type) -> List[dict]:
    """Full Weekly Schedule row set — every active task row plus every
    active leave row (already expanded per leave_rows_for_record), sorted
    per _row_sort_key. Callers pass only already-filtered active,
    in-week/overlapping-week rows; no additional filtering happens here."""
    rows = [task_row(event) for event in tasks]
    for record in leave_records:
        rows.extend(leave_rows_for_record(record, week_start, week_end))
    rows.sort(key=_row_sort_key)
    return rows


def _write_weekly_schedule_sheet(ws, rows: List[dict]) -> None:
    ws.append(WEEKLY_SCHEDULE_HEADERS)
    header_font = Font(bold=True)
    for col_index in range(1, len(WEEKLY_SCHEDULE_HEADERS) + 1):
        ws.cell(row=1, column=col_index).font = header_font

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([
            row["date"], row["day"], row["start_time"], row["end_time"],
            row["title"], row["priority"] or "", row["category"] or "",
            row["outcome"] or "", row["outcome_reason"] or "", row["notes"] or "",
            row["created_at"], row["updated_at"],
        ])
        excel_row = ws.max_row
        ws.cell(row=excel_row, column=1).number_format = "dd-mm-yyyy"
        if row["start_time"] is not None:
            ws.cell(row=excel_row, column=3).number_format = "hh:mm"
        if row["end_time"] is not None:
            ws.cell(row=excel_row, column=4).number_format = "hh:mm"
        if row["created_at"] is not None:
            ws.cell(row=excel_row, column=11).number_format = "dd-mm-yyyy hh:mm"
        if row["updated_at"] is not None:
            ws.cell(row=excel_row, column=12).number_format = "dd-mm-yyyy hh:mm"
        for col_index in (5, 9, 10):  # Title / Leave Type, Outcome Reason, Notes
            ws.cell(row=excel_row, column=col_index).alignment = wrap_alignment

    for col_index, width in enumerate(WEEKLY_SCHEDULE_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_weekly_summary_sheet(
    ws,
    member_key: str,
    member_label: str,
    week_start: date_type,
    week_end: date_type,
    aggregate: dict,
    leave_count: int,
    generated_at_colombo: datetime,
) -> None:
    bold = Font(bold=True)
    italic = Font(italic=True)

    ws.append([POINT_IN_TIME_NOTICE])
    ws.cell(row=1, column=1).font = italic
    ws.append([])

    fields = (
        ("Member", member_label),
        ("Week Start", week_start),
        ("Week End", week_end),
        ("Generated At", generated_at_colombo),
        ("Source", EXPORT_SOURCE_LABEL),
        ("Scheduled Task Count", aggregate["scheduled_count"]),
        ("Unscheduled Task Count", aggregate["unscheduled_count"]),
        ("Total Task Count", aggregate["total_count"]),
        ("Scheduled Duration", _format_duration(aggregate["scheduled_duration_minutes"])),
        ("Unscheduled Duration", _format_duration(aggregate["unscheduled_duration_minutes"])),
        ("Total Duration", _format_duration(aggregate["total_duration_minutes"])),
        ("Leave Count", leave_count),
    )

    ws.append(["Field", "Value"])
    header_row = ws.max_row
    ws.cell(row=header_row, column=1).font = bold
    ws.cell(row=header_row, column=2).font = bold

    for label, value in fields:
        ws.append([label, value])
        excel_row = ws.max_row
        ws.cell(row=excel_row, column=1).font = bold
        if label in ("Week Start", "Week End"):
            ws.cell(row=excel_row, column=2).number_format = "dd-mm-yyyy"
        elif label == "Generated At":
            ws.cell(row=excel_row, column=2).number_format = "dd-mm-yyyy hh:mm"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A" + str(header_row + 1)


def build_weekly_schedule_workbook(
    member_key: str,
    member_label: str,
    week_start: date_type,
    week_end: date_type,
    tasks,
    leave_records,
    aggregate: dict,
    leave_count: int,
    generated_at_utc: datetime,
) -> bytes:
    """Builds the full two-worksheet .xlsx and returns its bytes. `aggregate`
    must be the dict returned by member_schedules._aggregate_schedule_period
    for this exact (member_key, week_start, week_end) — this function never
    recomputes or alters it, only formats the values it already contains.
    Nothing here touches a database session; this is a pure bytes-in,
    bytes-out builder."""
    rows = build_weekly_schedule_rows(tasks, leave_records, week_start, week_end)

    workbook = Workbook()
    schedule_sheet = workbook.active
    schedule_sheet.title = "Weekly Schedule"
    _write_weekly_schedule_sheet(schedule_sheet, rows)

    summary_sheet = workbook.create_sheet("Weekly Summary")
    _write_weekly_summary_sheet(
        summary_sheet,
        member_key=member_key,
        member_label=member_label,
        week_start=week_start,
        week_end=week_end,
        aggregate=aggregate,
        leave_count=leave_count,
        generated_at_colombo=_to_colombo_naive(generated_at_utc),
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

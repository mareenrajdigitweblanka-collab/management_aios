"""Automated tests for the FINAL AUTHORITATIVE SAME-TASK MULTIPLE-TIME-PERIOD
RULE (2026-07-27) — supersedes and replaces
backend/tests/test_same_task_timed_untimed_rule.py, whose narrower
timed-vs-untimed-only coverage is now fully subsumed by the tests below
(that file predates the shared classify_same_task_conflict() classifier and
did not yet cover exact-duplicate, timed-overlap, or both-untimed
blocking).

For the same member, same normalized Task title, and same Task date:
  1. Different non-overlapping timed periods       -> ALLOW
  2. Adjacent timed periods                        -> ALLOW
  3. Exact same start and end time                 -> BLOCK (exact_task_duplicate)
  4. Any positive-duration timed overlap            -> BLOCK (same_task_time_overlap)
  5. Both Tasks untimed                             -> BLOCK (same_task_time_required)
  6. One timed and one untimed                      -> BLOCK (same_task_time_required)
Different Task titles may overlap and remain allowed.

Two layers, matching this repo's established conventions:
- Pure-function tests (no DB) for classify_same_task_conflict(), calling it
  directly with hand-built SameTaskOccurrence fixtures.
- Endpoint-level tests against a real, ephemeral in-memory SQLite database,
  calling the route functions directly with db=<a real Session> — the same
  pattern backend/tests/test_task_outcome_endpoint.py and
  backend/tests/test_weekly_schedule_export_endpoint.py already use (no
  live Neon/Postgres connection from this workstation; no httpx/TestClient
  dependency added). Required here (rather than a DB-free FakeSession) for
  Bulk Tasks' database-existing check and edit's exclude_key, both of which
  are real SQL filters a fake query object that ignores filter arguments
  cannot exercise.

Run with: python -m unittest backend.tests.test_same_task_multiple_time_period_rule
"""

import unittest
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO

from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.config import MEMBER_LABELS
from backend.database import Base
from backend.models import MemberLeaveRecord, MemberScheduleEvent
from backend.routers.member_schedules import (
    SAME_TASK_CONFLICT_BOTH_UNTIMED,
    SAME_TASK_CONFLICT_EXACT_DUPLICATE,
    SAME_TASK_CONFLICT_NONE,
    SAME_TASK_CONFLICT_TIME_OVERLAP,
    SAME_TASK_CONFLICT_TIMED_VS_UNTIMED,
    SameTaskOccurrence,
    classify_same_task_conflict,
    create_member_schedule_event,
    create_member_schedule_events_bulk,
    export_weekly_schedule,
    get_weekly_schedule_report,
    update_member_schedule_event,
    update_member_schedule_event_outcome,
)
from backend.schemas import (
    BulkTaskCreateRequest,
    BulkTaskRowIn,
    MemberScheduleEventCreate,
    MemberScheduleEventUpdate,
    TaskOutcomeUpdate,
)
from backend.time_utils import colombo_today


def _attach_schema(dbapi_conn, connection_record):
    dbapi_conn.execute("ATTACH DATABASE ':memory:' AS management_aios")


def _occ(key="existing", event_date=date(2099, 1, 5), title="Standup", start=None, end=None):
    return SameTaskOccurrence(key=key, event_date=event_date, title=title, start=start, end=end)


# ── Tests 1-12: pure-function classify_same_task_conflict() — no DB ────────
class ClassifySameTaskConflictTests(unittest.TestCase):
    DATE = date(2099, 1, 5)

    def test_1_distinct_periods_none(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 0), time(10, 0),
            [_occ(start=time(14, 0), end=time(15, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_NONE)

    def test_2_adjacent_periods_none(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(10, 0), time(11, 0),
            [_occ(start=time(9, 0), end=time(10, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_NONE)

    def test_3_exact_timed_match(self):
        result, conflict = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 0), time(10, 0),
            [_occ(start=time(9, 0), end=time(10, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_EXACT_DUPLICATE)
        self.assertIsNotNone(conflict)

    def test_4_partial_overlap(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 30), time(10, 30),
            [_occ(start=time(9, 0), end=time(10, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_TIME_OVERLAP)

    def test_5_candidate_contained_in_existing(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(10, 0), time(11, 0),
            [_occ(start=time(9, 0), end=time(12, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_TIME_OVERLAP)

    def test_6_candidate_contains_existing(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 0), time(12, 0),
            [_occ(start=time(10, 0), end=time(11, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_TIME_OVERLAP)

    def test_7_both_untimed(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", None, None, [_occ(start=None, end=None)],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_BOTH_UNTIMED)

    def test_8_existing_timed_candidate_untimed(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", None, None,
            [_occ(start=time(9, 0), end=time(10, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_TIMED_VS_UNTIMED)

    def test_9_existing_untimed_candidate_timed(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 0), time(10, 0),
            [_occ(start=None, end=None)],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_TIMED_VS_UNTIMED)

    def test_10_different_normalized_title_none(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 0), time(10, 0),
            [_occ(title="Retro", start=time(9, 0), end=time(10, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_NONE)

    def test_11_different_date_none(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 0), time(10, 0),
            [_occ(event_date=date(2099, 1, 6), start=time(9, 0), end=time(10, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_NONE)

    def test_12_excluded_edit_id_none(self):
        result, _ = classify_same_task_conflict(
            self.DATE, "Standup", time(9, 0), time(10, 0),
            [_occ(key="self-id", start=time(9, 0), end=time(10, 0))],
            exclude_key="self-id",
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_NONE)

    def test_title_normalization_case_and_whitespace(self):
        """Test 42 — existing title normalization behavior preserved."""
        result, _ = classify_same_task_conflict(
            self.DATE, "  STANDUP  ", time(9, 0), time(10, 0),
            [_occ(title="standup", start=time(9, 0), end=time(10, 0))],
        )
        self.assertEqual(result, SAME_TASK_CONFLICT_EXACT_DUPLICATE)


class SameTaskEndpointTestCase(unittest.TestCase):
    """Fresh, isolated in-memory SQLite database per test method."""

    def setUp(self):
        self.engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        event.listen(self.engine, "connect", _attach_schema)
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(bind=self.engine, autocommit=False, autoflush=False)

    def tearDown(self):
        self.engine.dispose()

    def make_session(self):
        return self.SessionLocal()

    def make_task(self, session, event_date, title="Standup", member_key="mayurika",
                  start_time=None, end_time=None, deleted=False):
        now = datetime.now(timezone.utc)
        task = MemberScheduleEvent(
            member_key=member_key, member_label=MEMBER_LABELS[member_key],
            event_date=event_date, title=title, category="Scheduled Task", priority="Medium",
            start_time=start_time, end_time=end_time,
            source_scope="dashboard_testing", is_official_truth=False,
            created_at=now, updated_at=now,
            deleted_at=now if deleted else None,
        )
        session.add(task)
        session.commit()
        session.refresh(task)
        return task

    def make_leave(self, session, leave_type, start_date, member_key="mayurika"):
        now = datetime.now(timezone.utc)
        leave = MemberLeaveRecord(
            member_key=member_key, member_label=MEMBER_LABELS[member_key],
            leave_type=leave_type, start_date=start_date, end_date=start_date,
            coordination_copy_only=True, policy_source_id="SRC-POLICY-001",
            effective_leave_minutes=540, created_at=now, updated_at=now,
        )
        session.add(leave)
        session.commit()
        return leave

    def create_payload(self, task_date, title="Standup", start=None, end=None):
        return MemberScheduleEventCreate(date=task_date, title=title, start=start, end=end)

    def row_count(self, session):
        return session.query(MemberScheduleEvent).count()

    def assert_conflict(self, response, expected_code):
        self.assertIsInstance(response, JSONResponse)
        self.assertEqual(response.status_code, 409)
        import json
        body = json.loads(response.body)
        self.assertEqual(body["error"], expected_code)
        return body


DATE = date(2099, 1, 5)


# ── Tests 13-20: Single Task create endpoint ────────────────────────────────
class SingleCreateTests(SameTaskEndpointTestCase):
    def test_13_distinct_periods_allowed(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(14, 0), time(15, 0)), db=session,
        )
        self.assertNotIsInstance(response, JSONResponse)
        self.assertEqual(self.row_count(session), 2)

    def test_14_adjacent_periods_allowed(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(10, 0), time(11, 0)), db=session,
        )
        self.assertNotIsInstance(response, JSONResponse)
        self.assertEqual(self.row_count(session), 2)

    def test_15_exact_duplicate_blocked_409(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        before = self.row_count(session)
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(9, 0), time(10, 0)), db=session,
        )
        self.assert_conflict(response, "exact_task_duplicate")
        self.assertEqual(self.row_count(session), before)

    def test_16_overlap_blocked_409(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        before = self.row_count(session)
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(9, 30), time(10, 30)), db=session,
        )
        self.assert_conflict(response, "same_task_time_overlap")
        self.assertEqual(self.row_count(session), before)

    def test_17_both_untimed_blocked_409(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup")
        before = self.row_count(session)
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup"), db=session,
        )
        self.assert_conflict(response, "same_task_time_required")
        self.assertEqual(self.row_count(session), before)

    def test_18_mixed_timed_untimed_blocked_409(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        before = self.row_count(session)
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup"), db=session,
        )
        self.assert_conflict(response, "same_task_time_required")
        self.assertEqual(self.row_count(session), before)

    def test_19_different_title_overlap_allowed(self):
        """Different-title timed overlap is no longer silently allowed
        (superseded 2026-07-27 by the LUNCH-BREAK AND DIFFERENT-TITLE
        TASK-OVERLAP CONFIRMATION feature — see
        test_schedule_advisory_confirmation.py) — it now requires an
        advisory confirmation before it is allowed. This test proves the
        hard same-title classifier itself still returns NONE (allowed, not
        hard-blocked) for a different title; confirmed end-to-end advisory
        behavior (zero-write first response, one-write confirmed retry) is
        covered in test_schedule_advisory_confirmation.py."""
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Retro", time(9, 0), time(10, 0)), db=session,
        )
        self.assertIsInstance(response, JSONResponse)
        self.assertEqual(response.status_code, 409)
        import json
        body = json.loads(response.body)
        self.assertEqual(body["error"], "schedule_confirmation_required")
        codes = {w["code"] for w in body["warnings"]}
        self.assertIn("different_task_time_overlap", codes)
        self.assertEqual(self.row_count(session), 1)

        confirmed = create_member_schedule_event(
            "mayurika",
            self.create_payload(DATE, "Retro", time(9, 0), time(10, 0)).model_copy(
                update={"confirmation_fingerprint": body["confirmation_fingerprint"]}
            ),
            db=session,
        )
        self.assertNotIsInstance(confirmed, JSONResponse)
        self.assertEqual(self.row_count(session), 2)

    def test_20_rejected_create_causes_no_write(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        before = self.row_count(session)
        create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(9, 0), time(10, 0)), db=session,
        )
        self.assertEqual(self.row_count(session), before)


# ── Tests 21-30: Bulk Tasks ──────────────────────────────────────────────────
class BulkTests(SameTaskEndpointTestCase):
    def test_21_internal_distinct_periods_allowed(self):
        session = self.make_session()
        rows = [
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
            BulkTaskRowIn(date=DATE, title="Standup", start=time(14, 0), end=time(15, 0)),
        ]
        result = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(result["status"], "created")
        self.assertEqual(result["created_count"], 2)

    def test_22_internal_adjacent_periods_allowed(self):
        session = self.make_session()
        rows = [
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
            BulkTaskRowIn(date=DATE, title="Standup", start=time(10, 0), end=time(11, 0)),
        ]
        result = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(result["status"], "created")
        self.assertEqual(result["created_count"], 2)

    def test_23_internal_exact_duplicate_blocks_whole_batch(self):
        session = self.make_session()
        rows = [
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
        ]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertIsInstance(response, JSONResponse)
        self.assertEqual(response.status_code, 422)
        import json
        body = json.loads(response.body)
        codes = {e["code"] for e in body["errors"]}
        self.assertIn("exact_task_duplicate", codes)
        self.assertEqual(self.row_count(session), 0)

    def test_24_internal_overlap_blocks_whole_batch(self):
        session = self.make_session()
        rows = [
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 30), end=time(10, 30)),
        ]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(response.status_code, 422)
        import json
        body = json.loads(response.body)
        codes = {e["code"] for e in body["errors"]}
        self.assertIn("same_task_time_overlap", codes)
        self.assertEqual(self.row_count(session), 0)

    def test_25_internal_both_untimed_blocks_whole_batch(self):
        session = self.make_session()
        rows = [
            BulkTaskRowIn(date=DATE, title="Standup"),
            BulkTaskRowIn(date=DATE, title="Standup"),
        ]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(response.status_code, 422)
        import json
        body = json.loads(response.body)
        codes = {e["code"] for e in body["errors"]}
        self.assertIn("same_task_time_required", codes)
        self.assertEqual(self.row_count(session), 0)

    def test_26_internal_timed_untimed_blocks_whole_batch(self):
        session = self.make_session()
        rows = [
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
            BulkTaskRowIn(date=DATE, title="Standup"),
        ]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(response.status_code, 422)
        import json
        body = json.loads(response.body)
        codes = {e["code"] for e in body["errors"]}
        self.assertIn("same_task_time_required", codes)
        self.assertEqual(self.row_count(session), 0)

    def test_27_existing_db_exact_duplicate_blocks_whole_batch(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        rows = [
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
            BulkTaskRowIn(date=DATE, title="Retro"),
        ]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(response.status_code, 422)
        import json
        body = json.loads(response.body)
        codes = {e["code"] for e in body["errors"]}
        self.assertIn("exact_task_duplicate", codes)
        self.assertEqual(self.row_count(session), 1)  # only the pre-existing row

    def test_28_existing_db_overlap_blocks_whole_batch(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        rows = [BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 30), end=time(10, 30))]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(response.status_code, 422)
        import json
        body = json.loads(response.body)
        codes = {e["code"] for e in body["errors"]}
        self.assertIn("same_task_time_overlap", codes)
        self.assertEqual(self.row_count(session), 1)

    def test_29_existing_db_untimed_conflict_blocks_whole_batch(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup")
        rows = [BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0))]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(response.status_code, 422)
        import json
        body = json.loads(response.body)
        codes = {e["code"] for e in body["errors"]}
        self.assertIn("same_task_time_required", codes)
        self.assertEqual(self.row_count(session), 1)

    def test_30_atomic_rollback_proven(self):
        """A batch with one valid row and one conflicting row inserts
        NEITHER — atomicity/no-partial-insertion proof."""
        session = self.make_session()
        rows = [
            BulkTaskRowIn(date=DATE, title="Valid task", start=time(8, 0), end=time(8, 30)),
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
            BulkTaskRowIn(date=DATE, title="Standup", start=time(9, 0), end=time(10, 0)),
        ]
        response = create_member_schedule_events_bulk("mayurika", BulkTaskCreateRequest(tasks=rows), db=session)
        self.assertEqual(response.status_code, 422)
        self.assertEqual(self.row_count(session), 0)


# ── Tests 31-38: Task editing ────────────────────────────────────────────────
class EditTests(SameTaskEndpointTestCase):
    def test_31_edit_to_distinct_period_allowed(self):
        session = self.make_session()
        editing_task = self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        response = update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(start=time(14, 0), end=time(15, 0)), db=session,
        )
        self.assertNotIsInstance(response, JSONResponse)
        self.assertEqual(response.start_time, time(14, 0))

    def test_32_edit_to_adjacent_period_allowed(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        editing_task = self.make_task(session, DATE, "Standup", start_time=time(14, 0), end_time=time(15, 0))
        response = update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(start=time(10, 0), end=time(11, 0)), db=session,
        )
        self.assertNotIsInstance(response, JSONResponse)

    def test_33_edit_to_exact_duplicate_blocked(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        editing_task = self.make_task(session, DATE, "Standup", start_time=time(14, 0), end_time=time(15, 0))
        response = update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(start=time(9, 0), end=time(10, 0)), db=session,
        )
        self.assert_conflict(response, "exact_task_duplicate")

    def test_34_edit_to_overlap_blocked(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        editing_task = self.make_task(session, DATE, "Standup", start_time=time(14, 0), end_time=time(15, 0))
        response = update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(start=time(9, 30), end=time(10, 30)), db=session,
        )
        self.assert_conflict(response, "same_task_time_overlap")

    def test_35_edit_to_both_untimed_conflict_blocked(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup")  # untimed
        editing_task = self.make_task(session, DATE, "Standup", start_time=time(14, 0), end_time=time(15, 0))
        response = update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(start=None, end=None), db=session,
        )
        self.assert_conflict(response, "same_task_time_required")

    def test_36_edit_to_mixed_timed_untimed_conflict_blocked(self):
        session = self.make_session()
        # The "other" Task must be UNTIMED — editing_task is becoming timed
        # (14:00-15:00), so only an existing untimed occurrence of the same
        # title/date conflicts with it (timed-vs-untimed).
        self.make_task(session, DATE, "Standup")
        editing_task = self.make_task(session, DATE, "Standup")  # also untimed, pre-edit
        response = update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(start=time(14, 0), end=time(15, 0)), db=session,
        )
        self.assert_conflict(response, "same_task_time_required")

    def test_37_self_row_excluded(self):
        """A no-op-ish edit (e.g. only notes change) must never be blocked
        by comparing the task against itself."""
        session = self.make_session()
        editing_task = self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        response = update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(notes="just a note"), db=session,
        )
        self.assertNotIsInstance(response, JSONResponse)
        self.assertEqual(response.notes, "just a note")

    def test_38_rejected_edit_causes_no_update(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        editing_task = self.make_task(session, DATE, "Standup", start_time=time(14, 0), end_time=time(15, 0))
        update_member_schedule_event(
            "mayurika", editing_task.id, MemberScheduleEventUpdate(start=time(9, 0), end=time(10, 0)), db=session,
        )
        session.expire_all()
        reloaded = session.get(MemberScheduleEvent, editing_task.id)
        self.assertEqual(reloaded.start_time, time(14, 0))
        self.assertEqual(reloaded.end_time, time(15, 0))


# ── Tests 39-47: additional regression coverage ──────────────────────────────
class RegressionTests(SameTaskEndpointTestCase):
    def test_39_soft_deleted_task_does_not_block(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", start_time=time(9, 0), end_time=time(10, 0), deleted=True)
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(9, 0), time(10, 0)), db=session,
        )
        self.assertNotIsInstance(response, JSONResponse)

    def test_40_different_member_does_not_block(self):
        session = self.make_session()
        self.make_task(session, DATE, "Standup", member_key="arun", start_time=time(9, 0), end_time=time(10, 0))
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(9, 0), time(10, 0)), db=session,
        )
        self.assertNotIsInstance(response, JSONResponse)

    def test_41_tamil_unicode_title_exact_duplicate_detected(self):
        session = self.make_session()
        title = "கூட்டம் தயாரிப்பு"
        self.make_task(session, DATE, title, start_time=time(9, 0), end_time=time(10, 0))
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, title, time(9, 0), time(10, 0)), db=session,
        )
        self.assert_conflict(response, "exact_task_duplicate")

    def test_42_title_normalization_behavior_preserved(self):
        session = self.make_session()
        self.make_task(session, DATE, "  Standup  ", start_time=time(9, 0), end_time=time(10, 0))
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "STANDUP", time(9, 0), time(10, 0)), db=session,
        )
        self.assert_conflict(response, "exact_task_duplicate")

    def test_43_task_leave_conflict_rules_unchanged(self):
        """A Task save blocked by active Leave must still return
        leave_conflict — the new same-task check must never suppress or
        replace the existing, unmodified Task/Leave overlap rule."""
        session = self.make_session()
        self.make_leave(session, "Full-Day", DATE)
        response = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Brand new task", time(9, 0), time(10, 0)), db=session,
        )
        self.assertIsInstance(response, JSONResponse)
        self.assertEqual(response.status_code, 409)
        import json
        body = json.loads(response.body)
        self.assertEqual(body["error"], "leave_conflict")

    def test_44_classification_unchanged_for_allowed_tasks(self):
        """Two allowed (non-conflicting) same-title Tasks must each still
        receive normal Scheduled/Unscheduled classification — unaffected
        by the new same-task pre-checks."""
        session = self.make_session()
        r1 = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(9, 0), time(10, 0)), db=session,
        )
        r2 = create_member_schedule_event(
            "mayurika", self.create_payload(DATE, "Standup", time(14, 0), time(15, 0)), db=session,
        )
        self.assertNotIsInstance(r1, JSONResponse)
        self.assertNotIsInstance(r2, JSONResponse)
        self.assertIn(r1.category, ("Scheduled Task", "Unscheduled Task"))
        self.assertIn(r2.category, ("Scheduled Task", "Unscheduled Task"))

    def test_45_task_outcomes_unchanged(self):
        """Task Outcome recording (a fully separate endpoint/column) must
        still work normally for a Task that passed the new same-task
        check at creation time."""
        session = self.make_session()
        today = colombo_today()
        task = self.make_task(session, today, "Standup", start_time=time(9, 0), end_time=time(10, 0))
        result = update_member_schedule_event_outcome(
            "mayurika", task.id, TaskOutcomeUpdate(outcome="Completed"), db=session,
        )
        self.assertNotIsInstance(result, JSONResponse)
        self.assertEqual(result.outcome, "Completed")

    def test_46_schedule_summary_unchanged(self):
        """Weekly Schedule Summary aggregation must still count every
        allowed (non-conflicting) same-title occurrence normally."""
        session = self.make_session()
        monday = DATE - timedelta(days=DATE.weekday())
        create_member_schedule_event(
            "mayurika", self.create_payload(monday, "Standup", time(9, 0), time(10, 0)), db=session,
        )
        create_member_schedule_event(
            "mayurika", self.create_payload(monday, "Standup", time(14, 0), time(15, 0)), db=session,
        )
        report = get_weekly_schedule_report("mayurika", week_start=monday, db=session)
        self.assertEqual(report.total_count, 2)

    def test_47_xlsx_export_includes_every_allowed_occurrence(self):
        session = self.make_session()
        monday = DATE - timedelta(days=DATE.weekday())
        create_member_schedule_event(
            "mayurika", self.create_payload(monday, "Standup", time(9, 0), time(10, 0)), db=session,
        )
        create_member_schedule_event(
            "mayurika", self.create_payload(monday, "Standup", time(14, 0), time(15, 0)), db=session,
        )
        result = export_weekly_schedule("mayurika", week_start=monday, db=session)
        self.assertNotIsInstance(result, JSONResponse)
        wb = load_workbook(BytesIO(result.body))
        rows = list(wb["Weekly Schedule"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 2)
        titles = [r[4] for r in rows]
        self.assertEqual(titles, ["Standup", "Standup"])


if __name__ == "__main__":
    unittest.main()

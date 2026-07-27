"""Automated tests for the weekly schedule .xlsx export
(member-weekly-schedule-xlsx-export, 2026-07-24).

Pure-function / in-memory-workbook tests only — no database connection is
required or attempted, matching this repo's established testing convention
(see backend/tests/test_schedule_duration_reports.py). Task/Leave rows are
built with SimpleNamespace fakes carrying only the attributes
backend/xlsx_export.py actually reads (mirrors the real ORM column names),
never a live MemberScheduleEvent/MemberLeaveRecord instance. Live download
behavior (actual endpoint response, Excel/Google Sheets opening) is verified
separately — see
validation/member-weekly-schedule-xlsx-export-check-2026-07-24.md.

Run with: python -m unittest backend.tests.test_weekly_schedule_xlsx_export
"""

import unittest
from datetime import date, datetime, time, timezone
from types import SimpleNamespace

from openpyxl import load_workbook

from backend.xlsx_export import (
    build_export_filename,
    build_weekly_schedule_rows,
    build_weekly_schedule_workbook,
    leave_rows_for_record,
    task_row,
)


def fake_task(
    event_date, title="Task", category="Scheduled Task", priority="Medium",
    start_time=None, end_time=None, outcome=None, outcome_reason=None,
    notes=None, created_at=None, updated_at=None,
):
    return SimpleNamespace(
        event_date=event_date, title=title, category=category, priority=priority,
        start_time=start_time, end_time=end_time, outcome=outcome,
        outcome_reason=outcome_reason, notes=notes,
        created_at=created_at or datetime(2026, 7, 20, 9, 0, tzinfo=timezone.utc),
        updated_at=updated_at or datetime(2026, 7, 20, 9, 0, tzinfo=timezone.utc),
    )


def fake_leave(
    leave_type, start_date, end_date=None, start_time=None, end_time=None,
    purpose=None, created_at=None, updated_at=None,
):
    return SimpleNamespace(
        leave_type=leave_type, start_date=start_date, end_date=end_date or start_date,
        start_time=start_time, end_time=end_time, purpose=purpose,
        created_at=created_at or datetime(2026, 7, 20, 9, 0, tzinfo=timezone.utc),
        updated_at=updated_at or datetime(2026, 7, 20, 9, 0, tzinfo=timezone.utc),
    )


class BuildExportFilenameTests(unittest.TestCase):
    def test_basic_filename(self):
        self.assertEqual(
            build_export_filename("mayurika", date(2026, 7, 27), date(2026, 8, 2)),
            "27-07-2026_to_02-08-2026_mayurika_weekly_schedule.xlsx",
        )

    def test_year_boundary_filename(self):
        self.assertEqual(
            build_export_filename("arun", date(2026, 12, 28), date(2027, 1, 3)),
            "28-12-2026_to_03-01-2027_arun_weekly_schedule.xlsx",
        )

    def test_no_spaces(self):
        name = build_export_filename("paraparan", date(2026, 1, 5), date(2026, 1, 11))
        self.assertNotIn(" ", name)


class TaskRowTests(unittest.TestCase):
    def test_pending_task_no_outcome(self):
        row = task_row(fake_task(date(2026, 7, 27)))
        self.assertEqual(row["outcome"], "Pending")
        self.assertIsNone(row["outcome_reason"])

    def test_uncompleted_task_includes_reason(self):
        t = fake_task(
            date(2026, 7, 20), outcome="Uncompleted", outcome_reason="Blocked by dependency",
        )
        row = task_row(t)
        self.assertEqual(row["outcome"], "Uncompleted")
        self.assertEqual(row["outcome_reason"], "Blocked by dependency")

    def test_completed_task_reason_blank(self):
        t = fake_task(date(2026, 7, 20), outcome="Completed", outcome_reason=None)
        row = task_row(t)
        self.assertEqual(row["outcome"], "Completed")
        self.assertIsNone(row["outcome_reason"])

    def test_no_response_for_past_untouched_task(self):
        row = task_row(fake_task(date(2020, 1, 1)))
        self.assertEqual(row["outcome"], "No response")

    def test_untimed_task_blank_start_end(self):
        row = task_row(fake_task(date(2026, 7, 27)))
        self.assertIsNone(row["start_time"])
        self.assertIsNone(row["end_time"])

    def test_day_name_matches_weekday(self):
        row = task_row(fake_task(date(2026, 7, 27)))  # a Monday
        self.assertEqual(row["day"], "Monday")


class LeaveRowsForRecordTests(unittest.TestCase):
    WEEK_START = date(2026, 7, 27)  # Monday
    WEEK_END = date(2026, 8, 2)     # Sunday

    def test_short_leave_single_row_with_time(self):
        lv = fake_leave("Short Leave", date(2026, 7, 28), start_time=time(9, 0), end_time=time(10, 0))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["title"], "Short Leave")
        self.assertEqual(rows[0]["start_time"], time(9, 0))
        self.assertEqual(rows[0]["priority"], None)
        self.assertEqual(rows[0]["category"], None)

    def test_half_day_first_label_and_blank_time(self):
        lv = fake_leave("Half-Day First", date(2026, 7, 29))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["title"], "First-Half Leave")
        self.assertIsNone(rows[0]["start_time"])

    def test_half_day_second_label(self):
        lv = fake_leave("Half-Day Second", date(2026, 7, 29))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        self.assertEqual(rows[0]["title"], "Second-Half Leave")

    def test_full_day_label(self):
        lv = fake_leave("Full-Day", date(2026, 7, 30))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        self.assertEqual(rows[0]["title"], "Full-Day Leave")

    def test_multi_day_expands_weekdays_only(self):
        # Wed 2026-07-29 through Tue 2026-08-04 (spans the week boundary) —
        # only Mon-Fri dates that also fall in [WEEK_START, WEEK_END]
        # appear: clamped to 2026-07-29..2026-08-02, so Wed/Thu/Fri only
        # (Sat 08-01/Sun 08-02 excluded as weekend, Mon 08-03/Tue 08-04
        # excluded as outside the requested week).
        lv = fake_leave("Multi-Day", date(2026, 7, 29), date(2026, 8, 4))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        dates = [r["date"] for r in rows]
        self.assertEqual(dates, [date(2026, 7, 29), date(2026, 7, 30), date(2026, 7, 31)])

    def test_multi_day_excludes_weekend(self):
        # Thu 2026-07-30 through Sun 2026-08-02 (both weekend dates fall
        # squarely inside the requested week) — Saturday/Sunday must not
        # appear even though they fall inside the leave's own date range.
        lv = fake_leave("Multi-Day", date(2026, 7, 30), date(2026, 8, 2))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        dates = [r["date"] for r in rows]
        self.assertNotIn(date(2026, 8, 1), dates)  # Saturday
        self.assertNotIn(date(2026, 8, 2), dates)  # Sunday
        self.assertIn(date(2026, 7, 30), dates)    # Thursday
        self.assertIn(date(2026, 7, 31), dates)    # Friday

    def test_multi_day_clamped_to_requested_week(self):
        # Leave spans well beyond the requested week on both sides — only
        # dates inside [WEEK_START, WEEK_END] may appear.
        lv = fake_leave("Multi-Day", date(2026, 7, 1), date(2026, 8, 31))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        for row in rows:
            self.assertTrue(self.WEEK_START <= row["date"] <= self.WEEK_END)

    def test_single_day_leave_outside_week_produces_no_rows(self):
        lv = fake_leave("Full-Day", date(2026, 6, 1))
        rows = leave_rows_for_record(lv, self.WEEK_START, self.WEEK_END)
        self.assertEqual(rows, [])


class BuildWeeklyScheduleRowsSortingTests(unittest.TestCase):
    def test_sorted_by_date_then_start_then_end_then_title(self):
        tasks = [
            fake_task(date(2026, 7, 27), title="Zebra task", start_time=time(9, 0), end_time=time(10, 0)),
            fake_task(date(2026, 7, 27), title="Alpha task"),  # untimed — sorts first
            fake_task(date(2026, 7, 26), title="Earlier day"),
        ]
        rows = build_weekly_schedule_rows(tasks, [], date(2026, 7, 20), date(2026, 7, 26))
        titles_in_order = [r["title"] for r in rows]
        self.assertEqual(titles_in_order[0], "Earlier day")
        # Untimed "Alpha task" (2026-07-27) sorts before timed "Zebra task"
        # (also 2026-07-27) — blank start/end sorts before any HH:MM string.
        self.assertEqual(titles_in_order[1], "Alpha task")
        self.assertEqual(titles_in_order[2], "Zebra task")

    def test_stable_for_equal_keys(self):
        tasks = [
            fake_task(date(2026, 7, 27), title="Same title"),
            fake_task(date(2026, 7, 27), title="Same title"),
        ]
        rows = build_weekly_schedule_rows(tasks, [], date(2026, 7, 27), date(2026, 7, 27))
        self.assertEqual(len(rows), 2)


class BuildWeeklyScheduleWorkbookTests(unittest.TestCase):
    """Round-trips a generated workbook through openpyxl.load_workbook to
    confirm structural correctness (valid OOXML, both required worksheet
    names present, header row intact, values readable) — the closest this
    DB-free test suite can get to Excel/Google Sheets opening it; the
    actual open in both applications is verified manually and recorded in
    validation/member-weekly-schedule-xlsx-export-check-2026-07-24.md."""

    def _sample_aggregate(self):
        return {
            "scheduled_count": 2, "unscheduled_count": 1, "total_count": 3,
            "scheduled_duration_minutes": 120, "unscheduled_duration_minutes": 30,
            "total_duration_minutes": 150,
        }

    def test_worksheet_names_and_header_row(self):
        tasks = [fake_task(date(2026, 7, 27), title="Write report")]
        leave_records = [fake_leave("Full-Day", date(2026, 7, 29))]
        workbook_bytes = build_weekly_schedule_workbook(
            member_key="mayurika", member_label="Mayurika — HR",
            week_start=date(2026, 7, 27), week_end=date(2026, 8, 2),
            tasks=tasks, leave_records=leave_records,
            aggregate=self._sample_aggregate(), leave_count=1,
            generated_at_utc=datetime(2026, 7, 27, 10, 0, tzinfo=timezone.utc),
        )
        from io import BytesIO
        wb = load_workbook(BytesIO(workbook_bytes))
        self.assertEqual(wb.sheetnames, ["Weekly Schedule", "Weekly Summary"])

        schedule_ws = wb["Weekly Schedule"]
        header = [cell.value for cell in schedule_ws[1]]
        self.assertEqual(header, [
            "Date", "Day", "Start Time", "End Time", "Title / Leave Type",
            "Priority", "Scheduled/Unscheduled Category", "Outcome",
            "Outcome Reason", "Notes", "Created At", "Updated At",
        ])
        self.assertEqual(schedule_ws.freeze_panes, "A2")
        self.assertIsNotNone(schedule_ws.auto_filter.ref)
        self.assertEqual(schedule_ws.max_row, 3)  # header + task + leave

    def test_weekly_summary_contains_point_in_time_notice_and_fields(self):
        workbook_bytes = build_weekly_schedule_workbook(
            member_key="suman", member_label="Suman — Recruiting Officer",
            week_start=date(2026, 7, 27), week_end=date(2026, 8, 2),
            tasks=[fake_task(date(2026, 7, 27))], leave_records=[],
            aggregate=self._sample_aggregate(), leave_count=0,
            generated_at_utc=datetime(2026, 7, 27, 10, 0, tzinfo=timezone.utc),
        )
        from io import BytesIO
        wb = load_workbook(BytesIO(workbook_bytes))
        summary_ws = wb["Weekly Summary"]
        self.assertIn("point-in-time export from Management AIOS", summary_ws["A1"].value)

        field_values = {row[0].value: row[1].value for row in summary_ws.iter_rows(min_row=4)}
        self.assertEqual(field_values["Member"], "Suman — Recruiting Officer")
        self.assertEqual(field_values["Source"], "Management AIOS")
        self.assertEqual(field_values["Scheduled Task Count"], 2)
        self.assertEqual(field_values["Leave Count"], 0)
        self.assertEqual(field_values["Scheduled Duration"], "2h 0m")

    def test_empty_tasks_and_leave_still_produces_valid_workbook(self):
        """The endpoint itself short-circuits before calling this builder
        when both are empty (see member_schedules.py
        export_weekly_schedule) — this only confirms the builder itself
        doesn't error if ever called with zero rows."""
        workbook_bytes = build_weekly_schedule_workbook(
            member_key="rajiv", member_label="Rajiv — Admin Manager",
            week_start=date(2026, 7, 27), week_end=date(2026, 8, 2),
            tasks=[], leave_records=[],
            aggregate={
                "scheduled_count": 0, "unscheduled_count": 0, "total_count": 0,
                "scheduled_duration_minutes": 0, "unscheduled_duration_minutes": 0,
                "total_duration_minutes": 0,
            },
            leave_count=0,
            generated_at_utc=datetime(2026, 7, 27, 10, 0, tzinfo=timezone.utc),
        )
        from io import BytesIO
        wb = load_workbook(BytesIO(workbook_bytes))
        self.assertEqual(wb["Weekly Schedule"].max_row, 1)  # header only


if __name__ == "__main__":
    unittest.main()

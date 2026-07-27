"""Endpoint-level tests for GET /api/member-schedules/{member_key}/reports/
weekly/export (member-weekly-schedule-xlsx-export, 2026-07-24).

Same ephemeral in-memory SQLite + direct route-function-call pattern as
backend/tests/test_task_outcome_endpoint.py (see that file's module
docstring for the full rationale — no live Neon/Postgres connection from
this workstation, no httpx/TestClient dependency added). The route function
(export_weekly_schedule) is called directly with db=<a real Session>,
bypassing FastAPI's own dependency-injection layer, which runs the exact
same function body a real HTTP request would.

Run with: python -m unittest backend.tests.test_weekly_schedule_export_endpoint
"""

import unittest
from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.config import MEMBER_LABELS
from backend.database import Base
from backend.models import MemberLeaveRecord, MemberScheduleEvent
from backend.routers.member_schedules import export_weekly_schedule


def _attach_schema(dbapi_conn, connection_record):
    dbapi_conn.execute("ATTACH DATABASE ':memory:' AS management_aios")


class WeeklyScheduleExportEndpointTestCase(unittest.TestCase):
    """Fresh, isolated in-memory SQLite database per test method — no
    cross-test data leakage, no shared state, no real network connection.
    No test in this file ever calls session.add()/commit() from inside the
    route function itself — only this fixture's own setup helpers seed
    rows, exactly mirroring how a real request would find pre-existing
    data already in Postgres."""

    def setUp(self):
        self.engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        event.listen(self.engine, "connect", _attach_schema)
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(bind=self.engine, autocommit=False, autoflush=False)

    def tearDown(self):
        self.engine.dispose()

    def make_session(self):
        return self.SessionLocal()

    def make_task(
        self, session, event_date, member_key="mayurika", title="Test task",
        start_time=None, end_time=None, deleted=False,
    ):
        now = datetime.now(timezone.utc)
        task = MemberScheduleEvent(
            member_key=member_key,
            member_label=MEMBER_LABELS[member_key],
            event_date=event_date,
            title=title,
            category="Scheduled Task",
            priority="Medium",
            start_time=start_time,
            end_time=end_time,
            source_scope="dashboard_testing",
            is_official_truth=False,
            created_at=now,
            updated_at=now,
            deleted_at=now if deleted else None,
        )
        session.add(task)
        session.commit()
        session.refresh(task)
        return task

    def make_leave(
        self, session, leave_type, start_date, end_date=None, member_key="mayurika", deleted=False,
    ):
        now = datetime.now(timezone.utc)
        leave = MemberLeaveRecord(
            member_key=member_key,
            member_label=MEMBER_LABELS[member_key],
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date or start_date,
            coordination_copy_only=True,
            policy_source_id="SRC-POLICY-001",
            effective_leave_minutes=540,
            created_at=now,
            updated_at=now,
            deleted_at=now if deleted else None,
        )
        session.add(leave)
        session.commit()
        session.refresh(leave)
        return leave

    def row_count(self, session, model):
        return session.query(model).count()


class EmptyWeekTests(WeeklyScheduleExportEndpointTestCase):
    def test_empty_week_returns_json_no_workbook(self):
        session = self.make_session()
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        self.assertIsInstance(result, JSONResponse)
        import json
        body = json.loads(result.body)
        self.assertTrue(body["empty"])
        self.assertEqual(body["week_start"], "2026-07-27")
        self.assertEqual(body["week_end"], "2026-08-02")

    def test_only_deleted_records_still_counts_as_empty(self):
        session = self.make_session()
        self.make_task(session, date(2026, 7, 27), deleted=True)
        self.make_leave(session, "Full-Day", date(2026, 7, 28), deleted=True)
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        self.assertIsInstance(result, JSONResponse)

    def test_other_member_records_do_not_prevent_empty_result(self):
        session = self.make_session()
        self.make_task(session, date(2026, 7, 27), member_key="arun")
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        self.assertIsInstance(result, JSONResponse)


class NonEmptyWeekTests(WeeklyScheduleExportEndpointTestCase):
    def test_task_only_week_returns_workbook_with_correct_filename(self):
        session = self.make_session()
        self.make_task(session, date(2026, 7, 27), title="Write report")
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        self.assertNotIsInstance(result, JSONResponse)
        self.assertEqual(
            result.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "27-07-2026_to_02-08-2026_mayurika_weekly_schedule.xlsx",
            result.headers["content-disposition"],
        )
        wb = load_workbook(BytesIO(result.body))
        self.assertEqual(wb.sheetnames, ["Weekly Schedule", "Weekly Summary"])
        rows = list(wb["Weekly Schedule"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][4], "Write report")  # Title / Leave Type column

    def test_leave_only_week(self):
        session = self.make_session()
        self.make_leave(session, "Full-Day", date(2026, 7, 29))
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        self.assertNotIsInstance(result, JSONResponse)
        wb = load_workbook(BytesIO(result.body))
        rows = list(wb["Weekly Schedule"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][4], "Full-Day Leave")

    def test_deleted_task_excluded(self):
        session = self.make_session()
        self.make_task(session, date(2026, 7, 27), title="Visible task")
        self.make_task(session, date(2026, 7, 27), title="Deleted task", deleted=True)
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        wb = load_workbook(BytesIO(result.body))
        titles = [r[4] for r in wb["Weekly Schedule"].iter_rows(min_row=2, values_only=True)]
        self.assertEqual(titles, ["Visible task"])

    def test_non_monday_week_start_normalizes_to_monday(self):
        session = self.make_session()
        self.make_task(session, date(2026, 7, 27), title="Any task")
        # Wednesday within the same week — must resolve to the same Monday.
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 29), db=session)
        self.assertIn(
            "27-07-2026_to_02-08-2026_mayurika_weekly_schedule.xlsx",
            result.headers["content-disposition"],
        )

    def test_month_boundary_week(self):
        session = self.make_session()
        self.make_task(session, date(2026, 7, 31), title="Month-end task")
        result = export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        self.assertIn(
            "27-07-2026_to_02-08-2026_mayurika_weekly_schedule.xlsx",
            result.headers["content-disposition"],
        )

    def test_year_boundary_week(self):
        session = self.make_session()
        self.make_task(session, date(2026, 12, 31), member_key="arun", title="Year-end task")
        result = export_weekly_schedule(member_key="arun", week_start=date(2026, 12, 28), db=session)
        self.assertIn(
            "28-12-2026_to_03-01-2027_arun_weekly_schedule.xlsx",
            result.headers["content-disposition"],
        )


class MemberIsolationTests(WeeklyScheduleExportEndpointTestCase):
    MEMBERS = ("mayurika", "suman", "arun", "rajiv", "paraparan")

    def test_each_member_only_sees_own_records(self):
        session = self.make_session()
        for member_key in self.MEMBERS:
            self.make_task(session, date(2026, 7, 27), member_key=member_key, title=member_key + "'s task")

        for member_key in self.MEMBERS:
            result = export_weekly_schedule(member_key=member_key, week_start=date(2026, 7, 27), db=session)
            self.assertIn(
                "_" + member_key + "_weekly_schedule.xlsx",
                result.headers["content-disposition"],
            )
            wb = load_workbook(BytesIO(result.body))
            titles = [r[4] for r in wb["Weekly Schedule"].iter_rows(min_row=2, values_only=True)]
            self.assertEqual(titles, [member_key + "'s task"])


class NoWriteProofTests(WeeklyScheduleExportEndpointTestCase):
    def test_export_never_changes_row_counts(self):
        session = self.make_session()
        self.make_task(session, date(2026, 7, 27))
        self.make_leave(session, "Full-Day", date(2026, 7, 29))
        before_tasks = self.row_count(session, MemberScheduleEvent)
        before_leave = self.row_count(session, MemberLeaveRecord)

        export_weekly_schedule(member_key="mayurika", week_start=date(2026, 7, 27), db=session)
        export_weekly_schedule(member_key="mayurika", week_start=date(2026, 6, 1), db=session)  # empty-week path too

        self.assertEqual(self.row_count(session, MemberScheduleEvent), before_tasks)
        self.assertEqual(self.row_count(session, MemberLeaveRecord), before_leave)


if __name__ == "__main__":
    unittest.main()
